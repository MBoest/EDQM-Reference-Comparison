from pathlib import Path
import datetime as dt
import pandas as pd
import openpyxl as ox
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import requests
import xml.etree.ElementTree as ET


pathxl = Path(".") / "input_file" / "example_reference_substances.xlsx"
path_edqm = Path(".") / "catalogues" / "web_catalog_XML.xml"
path_usp = Path(".") / "catalogues" / "usprefstd.csv"

url_edqm = "https://crs.edqm.eu/db/4DCGI/web_catalog_XML.xml"
url_usp = "https://static.usp.org/doc/referenceStandards/usprefstd.csv"

worksheet_name = "Reference Substances"


def create_df(ws) -> pd.DataFrame:
    if not ws:
        raise ValueError("No worksheet passed")
    
    start_row = 3

    data = {
        "Substance Specific Code": [
            ws.cell(row=i, column=1).value for i in range(start_row, ws.max_row + 1)
        ],
        "Substance Received Year": [
            ws.cell(row=i, column=2).value for i in range(start_row, ws.max_row + 1)
        ],
        "Substance Received Index": [
            ws.cell(row=i, column=3).value for i in range(start_row, ws.max_row + 1)
        ],
        "Substance": [
            ws.cell(row=i, column=4).value for i in range(start_row, ws.max_row + 1)
        ],
        "Article No.": [
            ws.cell(row=i, column=5).value for i in range(start_row, ws.max_row + 1)
        ],
        "In Stock": [
            ws.cell(row=i, column=6).value for i in range(start_row, ws.max_row + 1)
        ],
        "Category": [
            ws.cell(row=i, column=7).value for i in range(start_row, ws.max_row + 1)
        ],
        "Supplier Batch No.": [
            ws.cell(row=i, column=8).value for i in range(start_row, ws.max_row + 1)
        ],
    }

    return pd.DataFrame(data)

def filter_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.sort_values(by="Substance Specific Code")

    df_refs_in_stock = df[
        df["In Stock"].str.contains("yes", case=False, na=False)
        & 
            df["Category"].str.contains("CRS|USP", case=False, na=False)
    ].copy()

    df_refs_in_stock["In Stock"] = df_refs_in_stock["In Stock"].astype("string").str.strip()
    df_refs_in_stock["Category"] = df_refs_in_stock["Category"].astype("string").str.strip()
    df_refs_in_stock["Article No."] = df_refs_in_stock["Article No."].astype("string").str.strip()

    # Only clean letters from edqm codes
    mask_edqm = df_refs_in_stock["Category"].str.contains(
    "CRS",
    case=False,
    na=False
    )
    
    df_refs_in_stock.loc[mask_edqm, "Supplier Batch No."] = (
        df_refs_in_stock.loc[mask_edqm, "Supplier Batch No."]
        .astype("string")
        .str.replace(r"[A-Za-z]", "", regex=True)
        .str.lstrip("0")
        .str.split(".")
        .str[0]
    )

    return df_refs_in_stock

def download_catalogue(url: str, destination: Path) -> None:
    try:
        response = requests.get(url, timeout=90)
        response.raise_for_status()
    except requests.RequestException as e:
        raise ValueError(f"Download failed: {e}") from e

    with open(destination, "wb") as file:
        file.write(response.content)

def read_xml_file(pathxml: Path) -> pd.DataFrame:
    tree = ET.parse(pathxml)
    root = tree.getroot()

    edqm_references = []

    for ref in root.findall("Reference"):
        order_code = ref.get("Order_Code")
        batch_edqm = ref.find("Batch_No")
        if batch_edqm is None or batch_edqm.text is None:
          continue
        batch_edqm = batch_edqm.text

        edqm_references.append({
            "Article No.": order_code,
            "Current Batch": batch_edqm,
            })
    
    return pd.DataFrame(edqm_references)

def read_csv_file(pathcsv: Path) -> pd.DataFrame:
    df = pd.read_csv(pathcsv)
    df = df[["Catalog #", "Current Lot"]]
    df = df.rename(columns={"Catalog #": "Article No.", "Current Lot": "Current Batch"})

    return df

def compare_batchnumbers(own_references: pd.DataFrame, current_catalogue: pd.DataFrame) -> pd.DataFrame:
    shared_codes = pd.merge(own_references, current_catalogue, on="Article No.", how="left",
    )

    return shared_codes[
        shared_codes["Supplier Batch No."] != shared_codes["Current Batch"]
    ]

def write_edqm_check_to_workbook(wb: ox.Workbook, standards_to_check_df: pd.DataFrame, file_path: Path) -> None:
    sheet_name = f"EDQM_check_{dt.date.today():%Y_%m_%d}"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    df = standards_to_check_df.copy()

    df.columns = [str(col) for col in df.columns]

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2:
        wb.save(file_path)
        return

    table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table_name = f"EDQMCheck_{dt.datetime.now():%Y%m%d_%H%M%S}"

    table = Table(
        displayName=table_name,
        ref=table_ref,
    )

    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 40)

    wb.save(file_path)


if __name__ == "__main__":
    wb = ox.load_workbook(pathxl)

    ws = wb[worksheet_name]

    own_references_df = create_df(ws)

    if own_references_df.empty:
        raise ValueError("Empty DataFrame")

    own_references_df_filtered = filter_df(own_references_df)

    download_catalogue(url_edqm, path_edqm)
    download_catalogue(url_usp, path_usp)

    edqm_references_df = read_xml_file(path_edqm)
    usp_references_df = read_csv_file(path_usp)

    catalogues = [edqm_references_df, usp_references_df]
    combined_catalogues = pd.concat(catalogues)

    standards_to_check_df = compare_batchnumbers(own_references=own_references_df_filtered, current_catalogue=combined_catalogues)

    write_edqm_check_to_workbook(
        wb=wb,
        standards_to_check_df=standards_to_check_df,
        file_path=pathxl,
    )